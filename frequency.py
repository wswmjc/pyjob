#!/usr/bin/env python
"""
Minimal Example
===============

Generating a square wordcloud from the US constitution using default arguments.
"""

from os import path
from scipy.misc import imread
from wordcloud import WordCloud
from openpyxl import load_workbook
import matplotlib.pyplot as plt


d = path.dirname(__file__)

def get_txt_dict(xls_path):
    '''
        从excel里面读取数据，并返回dict
    '''
    try:
        wb = load_workbook(xls_path)
        ws = wb.active
        fre_dict = dict()
        row_gen = list(ws.iter_rows())
        for row in row_gen[1:]:
            fre_dict[row[0].value] = row[2].value
        return fre_dict
    except:
        pass
    return None


def show_img(word_cloud: WordCloud):
    '''
        展示词云图片
    '''
    plt.imshow(word_cloud, interpolation='bilinear')
    plt.axis("off")
    image = word_cloud.to_image()
    image.show()


def get_word_cloud(word_path: str = None,
                   word_dict: dict = None,
                   font_path: str = 'msyh.ttf',
                   background_color: str = 'black',
                   background_img_path: str = None):
    r'''
        word_path: string
            词云输入，可以传入excel文档路径： “SUPOR苏泊尔 的WordCloud_v1.xlsx”
        word_dict: dict
            处理过的词字典，{词：频率，九阳：0.999，...}
        font_path: str
            字体路径，用于设置展示词云的字体
        background_color: str
            背景颜色，基本颜色英文单词，“white”,“black”,....
        background_img_path: str
            词云背景形状图片路径，会根据改图片生成对应形状的词云
    '''
    if not (word_path or word_dict):
        raise Exception('word_path or word_dict 至少有一个不为None')
        
    background_img = None
    if background_img_path is not None:
        background_img = imread(path.join(d, background_img_path))
    
    word = word_dict if word_dict is not None else get_txt_dict(word_path)

    wordcloud = WordCloud(font_path=font_path,
                          background_color=background_color,
                          max_words=2000,  # 词云显示的最大词数
                          mask=background_img,  # 设置背景图片
                          max_font_size=100,  # 字体最大值
                          random_state=42,
                          width=100, height=260, margin=1 # 设置图片默认的大小,但是如果使用背景图片的话,那么保存的图片大小将会按照其大小保存,margin为词语边缘距离
                          ).generate_from_frequencies(word)
    show_img(wordcloud)


get_word_cloud(word_path='SUPOR苏泊尔 的WordCloud_v1.xlsx',background_img_path='9yang.jpg',background_color='white')
